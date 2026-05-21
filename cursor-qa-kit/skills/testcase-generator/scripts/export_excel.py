"""
export_excel.py
---------------
将 AI 生成的测试用例同时导出为 xlsx 和 mm（FreeMind）格式。

规则：
- # 一级标题  → 「模块」列及标题前缀（主模块名）
- ## 二级标题 → 「子模块」列的值
- 全部用例导出到**同一个 Sheet**；「模块」列中连续相同主模块的行做**纵向合并单元格**
- 标题列自动加前缀：`【主模块名】` + 原标题

依赖：
    pip install openpyxl

用法：
    python scripts/export_excel.py <markdown_file> [output_dir]

示例：
    python scripts/export_excel.py testcases/test_手动创建Skill-testcases.md testcases/
"""

import sys
import os
import re
import xml.etree.ElementTree as ET
from xml.dom import minidom

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("请先安装依赖：pip install openpyxl")
    sys.exit(1)


# 原始字段（不含子模块列）
FIELDS = ["编号", "标题", "测试点", "用例类型", "前置条件", "步骤",
          "预期结果", "实际结果", "执行状态", "优先级"]

# Excel 表头（编号后：模块、子模块；标题含【模块】前缀由脚本写入）
HEADERS = ["编号", "模块", "子模块", "标题", "测试点", "用例类型", "前置条件", "步骤",
           "预期结果", "实际结果", "执行状态", "优先级"]

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBMODULE_FILL = PatternFill("solid", fgColor="D9E1F2")
MODULE_FILL = PatternFill("solid", fgColor="E2EFDA")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def parse_markdown_tables(filepath: str) -> dict:
    """
    解析 Markdown 文件，返回结构：
    {
      "主模块名": [
        ["编号", "子模块名", "标题", "测试点", ...],  # 每行 11 列（含子模块）
        ...
      ]
    }
    导出 xlsx 时再展开为「模块」列，并对标题加「【主模块】」前缀。
    """
    result = {}
    current_main = "默认模块"
    current_sub = ""
    in_table = False

    with open(filepath, encoding="utf-8") as f:
        lines = f.readlines()

    for line in lines:
        line = line.rstrip()

        # 一级标题 → 主模块（写入「模块」列及标题前缀）
        if re.match(r"^# [^#]", line):
            current_main = line.lstrip("# ").strip()
            current_sub = ""
            in_table = False

        # 二级标题 → 子模块（列值）
        elif re.match(r"^## [^#]", line):
            current_sub = line.lstrip("# ").strip()
            in_table = False

        # 表格行
        elif "|" in line:
            cells = [c.strip() for c in line.split("|")[1:-1]]
            # 跳过分隔行
            if all(re.match(r"^[-:]+$", c) for c in cells if c):
                continue
            # 跳过表头行（原始字段表头 或 含子模块的表头）
            if cells == FIELDS or cells == HEADERS:
                in_table = True
                continue
            # 匹配原始字段数量的数据行
            if len(cells) == len(FIELDS):
                in_table = True
                # 内部行：[编号, 子模块, 标题, ...]（主模块在 dict 的 key 中）
                row_with_sub = [cells[0], current_sub] + cells[1:]
                result.setdefault(current_main, []).append(row_with_sub)

    return result


def _merge_consecutive_same_column(ws, col_idx: int, first_data_row: int):
    """将指定列中连续相同非空取值的行纵向合并（值保留在合并区域左上角）。"""
    max_r = ws.max_row
    r = first_data_row
    while r <= max_r:
        v = ws.cell(row=r, column=col_idx).value
        if v is None or (isinstance(v, str) and not str(v).strip()):
            r += 1
            continue
        r2 = r
        while r2 + 1 <= max_r and ws.cell(row=r2 + 1, column=col_idx).value == v:
            r2 += 1
        if r2 > r:
            ws.merge_cells(
                start_row=r, start_column=col_idx, end_row=r2, end_column=col_idx
            )
            top = ws.cell(row=r, column=col_idx)
            top.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            top.border = THIN_BORDER
        r = r2 + 1


def export_xlsx(tables: dict, output_path: str, sheet_title: str = "测试用例"):
    """导出 xlsx：全部用例写入同一 Sheet；模块列纵向合并；标题加【模块】前缀。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = re.sub(r"[\\/*?:\[\]]", "-", sheet_title)[:31] or "测试用例"

    flat_rows = []
    for module, rows in tables.items():
        for row in rows:
            flat_rows.append((module, row))

    # 写表头
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    prev_sub_key = None
    for main_module, row in flat_rows:
        sub_val = row[1] if len(row) > 1 else ""
        title_orig = row[2] if len(row) > 2 else ""
        title_disp = (
            f"【{main_module}】{title_orig}"
            if title_orig
            else f"【{main_module}】"
        )
        excel_row = [row[0], main_module, sub_val, title_disp] + row[3:]
        ws.append(excel_row)
        current_row = ws[ws.max_row]

        for cell in current_row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = THIN_BORDER

        mod_cell = current_row[1]  # 列「模块」
        mod_cell.fill = MODULE_FILL
        mod_cell.font = Font(bold=True)

        sub_cell = current_row[2]
        sub_key = (main_module, sub_val)
        if sub_key != prev_sub_key:
            sub_cell.fill = SUBMODULE_FILL
            sub_cell.font = Font(bold=True)
            prev_sub_key = sub_key
        else:
            sub_cell.value = ""

    if ws.max_row >= 2:
        _merge_consecutive_same_column(ws, col_idx=2, first_data_row=2)

    col_widths = {
        "A": 10,   # 编号
        "B": 14,   # 模块
        "C": 16,   # 子模块
        "D": 28,   # 标题（含【模块】前缀）
        "E": 16,   # 测试点
        "F": 14,   # 用例类型
        "G": 20,   # 前置条件
        "H": 40,   # 步骤
        "I": 40,   # 预期结果
        "J": 14,   # 实际结果
        "K": 12,   # 执行状态
        "L": 10,   # 优先级
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"[xlsx] 已保存：{output_path}")


def export_mm(tables: dict, output_path: str):
    """导出 FreeMind .mm 文件，使用文件第一个主模块名作为顶节点"""
    root_node = ET.Element("map", version="1.0.1")

    # 顶节点使用第一个主模块名（或文件名）
    top_text = next(iter(tables), "测试用例")
    root_mind = ET.SubElement(root_node, "node", TEXT=top_text)

    # 按子模块分组
    sub_groups: dict = {}
    for rows in tables.values():
        for row in rows:
            sub = row[1] if len(row) > 1 else "默认"
            sub_groups.setdefault(sub, []).append(row)

    for sub, rows in sub_groups.items():
        sub_node = ET.SubElement(root_mind, "node", TEXT=sub, POSITION="right")
        for row in rows:
            # row: [编号, 子模块, 标题, 测试点, 用例类型, 前置条件, 步骤, 预期结果, ...]
            title = row[2] if len(row) > 2 else "无标题"
            priority = row[10] if len(row) > 10 else ""  # 内部行仍为 11 列
            case_node = ET.SubElement(sub_node, "node",
                                      TEXT=f"[{row[0]}] {title}｜{priority}")
            labels = ["测试点", "用例类型", "前置条件", "步骤", "预期结果"]
            indices = [3, 4, 5, 6, 7]
            for label, idx in zip(labels, indices):
                if idx < len(row) and row[idx]:
                    ET.SubElement(case_node, "node", TEXT=f"{label}：{row[idx]}")

    xml_str = minidom.parseString(
        ET.tostring(root_node, encoding="unicode")
    ).toprettyxml(indent="  ")
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(xml_str)
    print(f"[mm]   已保存：{output_path}")


def main():
    if len(sys.argv) < 2:
        print("用法：python scripts/export_excel.py <markdown_file> [output_dir]")
        sys.exit(1)

    md_file = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else os.path.dirname(md_file)
    os.makedirs(output_dir, exist_ok=True)

    base_name = os.path.splitext(os.path.basename(md_file))[0]
    xlsx_path = os.path.join(output_dir, f"{base_name}.xlsx")
    mm_path = os.path.join(output_dir, f"{base_name}.mm")

    tables = parse_markdown_tables(md_file)
    if not tables:
        print("未解析到任何表格数据，请检查 Markdown 文件格式。")
        sys.exit(1)

    safe_sheet = re.sub(r"[\\/*?:\[\]]", "-", base_name)[:31] or "测试用例"
    export_xlsx(tables, xlsx_path, sheet_title=safe_sheet)
    export_mm(tables, mm_path)
    print("导出完成。")


if __name__ == "__main__":
    main()
